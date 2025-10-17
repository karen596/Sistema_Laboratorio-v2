# -*- coding: utf-8 -*-
"""
Módulo de Generación de Reportes
Sistema de Laboratorios - Centro Minero SENA
Genera reportes en formato PDF y Excel
"""

from datetime import datetime
from io import BytesIO
import os

# Importaciones para PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Importaciones para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """Clase para generar reportes en PDF y Excel"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Configurar estilos personalizados para PDF"""
        # Estilo para título principal
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e5128'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        # Estilo para subtítulos
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2d6a4f'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        ))
        
        # Estilo para texto normal
        self.styles.add(ParagraphStyle(
            name='CustomBody',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            fontName='Helvetica'
        ))
    
    def generar_pdf_estadisticas(self, data, fecha_inicio=None, fecha_fin=None):
        """
        Generar reporte PDF con estadísticas generales
        
        Args:
            data: Diccionario con los datos del reporte
            fecha_inicio: Fecha de inicio del reporte
            fecha_fin: Fecha de fin del reporte
        
        Returns:
            BytesIO: Buffer con el contenido del PDF
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        # Encabezado
        story.append(Paragraph("Sistema de Laboratorios", self.styles['CustomTitle']))
        story.append(Paragraph("Centro Minero SENA", self.styles['CustomHeading']))
        story.append(Spacer(1, 0.2*inch))
        
        # Información del reporte
        fecha_generacion = datetime.now().strftime("%d/%m/%Y %H:%M")
        story.append(Paragraph(f"<b>Fecha de generación:</b> {fecha_generacion}", self.styles['CustomBody']))
        
        if fecha_inicio and fecha_fin:
            story.append(Paragraph(f"<b>Período:</b> {fecha_inicio} - {fecha_fin}", self.styles['CustomBody']))
        
        story.append(Spacer(1, 0.3*inch))
        
        # Estadísticas Generales
        story.append(Paragraph("Estadísticas Generales", self.styles['CustomHeading']))
        
        estadisticas_data = [
            ['Métrica', 'Valor'],
            ['Total de Equipos', str(data.get('total_equipos', 0))],
            ['Equipos Activos', str(data.get('equipos_activos', 0))],
            ['Total de Usuarios', str(data.get('total_usuarios', 0))],
            ['Total de Reservas', str(data.get('total_reservas', 0))],
            ['Reservas Activas', str(data.get('reservas_activas', 0))],
            ['Items en Inventario', str(data.get('total_items', 0))],
            ['Items con Stock Bajo', str(data.get('items_stock_bajo', 0))],
        ]
        
        tabla_estadisticas = Table(estadisticas_data, colWidths=[3.5*inch, 2*inch])
        tabla_estadisticas.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d6a4f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(tabla_estadisticas)
        story.append(Spacer(1, 0.3*inch))
        
        # Equipos más utilizados
        if 'equipos_mas_usados' in data and data['equipos_mas_usados']:
            story.append(Paragraph("Equipos Más Utilizados", self.styles['CustomHeading']))
            
            equipos_data = [['Equipo', 'Número de Usos']]
            for equipo in data['equipos_mas_usados'][:10]:
                equipos_data.append([
                    equipo.get('nombre', 'N/A'),
                    str(equipo.get('usos', 0))
                ])
            
            tabla_equipos = Table(equipos_data, colWidths=[4*inch, 1.5*inch])
            tabla_equipos.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d6a4f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(tabla_equipos)
            story.append(Spacer(1, 0.3*inch))
        
        # Usuarios más activos
        if 'usuarios_activos' in data and data['usuarios_activos']:
            story.append(Paragraph("Usuarios Más Activos", self.styles['CustomHeading']))
            
            usuarios_data = [['Usuario', 'Tipo', 'Actividad']]
            for usuario in data['usuarios_activos'][:10]:
                usuarios_data.append([
                    usuario.get('nombre', 'N/A'),
                    usuario.get('tipo', 'N/A'),
                    str(usuario.get('comandos', 0))
                ])
            
            tabla_usuarios = Table(usuarios_data, colWidths=[3*inch, 1.5*inch, 1*inch])
            tabla_usuarios.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d6a4f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(tabla_usuarios)
            story.append(Spacer(1, 0.3*inch))
        
        # Inventario con stock bajo
        if 'inventario_bajo' in data and data['inventario_bajo']:
            story.append(PageBreak())
            story.append(Paragraph("Inventario con Stock Bajo", self.styles['CustomHeading']))
            
            inventario_data = [['Item', 'Categoría', 'Stock Actual', 'Stock Mínimo']]
            for item in data['inventario_bajo']:
                inventario_data.append([
                    item.get('nombre', 'N/A'),
                    item.get('categoria', 'N/A'),
                    str(item.get('cantidad_actual', 0)),
                    str(item.get('cantidad_minima', 0))
                ])
            
            tabla_inventario = Table(inventario_data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1*inch])
            tabla_inventario.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d97706')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(tabla_inventario)
        
        # Pie de página
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph(
            f"Reporte generado automáticamente por el Sistema de Laboratorios - Centro Minero SENA",
            self.styles['CustomBody']
        ))
        
        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generar_excel_estadisticas(self, data, fecha_inicio=None, fecha_fin=None):
        """
        Generar reporte Excel con estadísticas generales
        
        Args:
            data: Diccionario con los datos del reporte
            fecha_inicio: Fecha de inicio del reporte
            fecha_fin: Fecha de fin del reporte
        
        Returns:
            BytesIO: Buffer con el contenido del Excel
        """
        buffer = BytesIO()
        wb = Workbook()
        
        # Estilos
        header_fill = PatternFill(start_color="2d6a4f", end_color="2d6a4f", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16, color="1e5128")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Hoja 1: Resumen General
        ws1 = wb.active
        ws1.title = "Resumen General"
        
        # Título
        ws1['A1'] = "Sistema de Laboratorios - Centro Minero SENA"
        ws1['A1'].font = title_font
        ws1.merge_cells('A1:D1')
        ws1['A1'].alignment = Alignment(horizontal='center')
        
        # Información del reporte
        ws1['A3'] = "Fecha de generación:"
        ws1['B3'] = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        if fecha_inicio and fecha_fin:
            ws1['A4'] = "Período:"
            ws1['B4'] = f"{fecha_inicio} - {fecha_fin}"
        
        # Estadísticas generales
        row = 6
        ws1[f'A{row}'] = "Métrica"
        ws1[f'B{row}'] = "Valor"
        ws1[f'A{row}'].fill = header_fill
        ws1[f'B{row}'].fill = header_fill
        ws1[f'A{row}'].font = header_font
        ws1[f'B{row}'].font = header_font
        
        estadisticas = [
            ('Total de Equipos', data.get('total_equipos', 0)),
            ('Equipos Activos', data.get('equipos_activos', 0)),
            ('Total de Usuarios', data.get('total_usuarios', 0)),
            ('Total de Reservas', data.get('total_reservas', 0)),
            ('Reservas Activas', data.get('reservas_activas', 0)),
            ('Items en Inventario', data.get('total_items', 0)),
            ('Items con Stock Bajo', data.get('items_stock_bajo', 0)),
        ]
        
        row += 1
        for metrica, valor in estadisticas:
            ws1[f'A{row}'] = metrica
            ws1[f'B{row}'] = valor
            ws1[f'A{row}'].border = border
            ws1[f'B{row}'].border = border
            row += 1
        
        # Ajustar ancho de columnas
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 15
        
        # Hoja 2: Equipos Más Utilizados
        if 'equipos_mas_usados' in data and data['equipos_mas_usados']:
            ws2 = wb.create_sheet("Equipos Más Usados")
            
            ws2['A1'] = "Equipos Más Utilizados"
            ws2['A1'].font = title_font
            ws2.merge_cells('A1:B1')
            
            ws2['A3'] = "Equipo"
            ws2['B3'] = "Número de Usos"
            ws2['A3'].fill = header_fill
            ws2['B3'].fill = header_fill
            ws2['A3'].font = header_font
            ws2['B3'].font = header_font
            
            row = 4
            for equipo in data['equipos_mas_usados']:
                ws2[f'A{row}'] = equipo.get('nombre', 'N/A')
                ws2[f'B{row}'] = equipo.get('usos', 0)
                ws2[f'A{row}'].border = border
                ws2[f'B{row}'].border = border
                row += 1
            
            ws2.column_dimensions['A'].width = 40
            ws2.column_dimensions['B'].width = 15
        
        # Hoja 3: Usuarios Más Activos
        if 'usuarios_activos' in data and data['usuarios_activos']:
            ws3 = wb.create_sheet("Usuarios Más Activos")
            
            ws3['A1'] = "Usuarios Más Activos"
            ws3['A1'].font = title_font
            ws3.merge_cells('A1:C1')
            
            ws3['A3'] = "Usuario"
            ws3['B3'] = "Tipo"
            ws3['C3'] = "Actividad"
            for col in ['A3', 'B3', 'C3']:
                ws3[col].fill = header_fill
                ws3[col].font = header_font
            
            row = 4
            for usuario in data['usuarios_activos']:
                ws3[f'A{row}'] = usuario.get('nombre', 'N/A')
                ws3[f'B{row}'] = usuario.get('tipo', 'N/A')
                ws3[f'C{row}'] = usuario.get('comandos', 0)
                for col in ['A', 'B', 'C']:
                    ws3[f'{col}{row}'].border = border
                row += 1
            
            ws3.column_dimensions['A'].width = 30
            ws3.column_dimensions['B'].width = 20
            ws3.column_dimensions['C'].width = 15
        
        # Hoja 4: Inventario con Stock Bajo
        if 'inventario_bajo' in data and data['inventario_bajo']:
            ws4 = wb.create_sheet("Inventario Crítico")
            
            ws4['A1'] = "Inventario con Stock Bajo"
            ws4['A1'].font = title_font
            ws4.merge_cells('A1:D1')
            
            ws4['A3'] = "Item"
            ws4['B3'] = "Categoría"
            ws4['C3'] = "Stock Actual"
            ws4['D3'] = "Stock Mínimo"
            for col in ['A3', 'B3', 'C3', 'D3']:
                ws4[col].fill = PatternFill(start_color="d97706", end_color="d97706", fill_type="solid")
                ws4[col].font = header_font
            
            row = 4
            for item in data['inventario_bajo']:
                ws4[f'A{row}'] = item.get('nombre', 'N/A')
                ws4[f'B{row}'] = item.get('categoria', 'N/A')
                ws4[f'C{row}'] = item.get('cantidad_actual', 0)
                ws4[f'D{row}'] = item.get('cantidad_minima', 0)
                for col in ['A', 'B', 'C', 'D']:
                    ws4[f'{col}{row}'].border = border
                row += 1
            
            ws4.column_dimensions['A'].width = 30
            ws4.column_dimensions['B'].width = 20
            ws4.column_dimensions['C'].width = 15
            ws4.column_dimensions['D'].width = 15
        
        # Guardar en buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer


# Instancia global del generador
report_generator = ReportGenerator()
